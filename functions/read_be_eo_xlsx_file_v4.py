import pandas as pd
from extensions import extensions
from initial_values.initial_values import be_data_columns_to_master_columns, year_dict
from datetime import datetime
from initial_values.initial_values import sap_user_status_cons_status_list, be_data_cons_status_list, sap_system_status_ban_list, operaton_status_translation, master_data_to_ru_columns, month_dict
import sqlite3
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl

db = extensions.db

date_time_plug = '1.1.2199'
date_time_plug = datetime.strptime(date_time_plug, '%d.%m.%Y')

# наименование итераций
iterations = {
  "Дата вывода 0 - итерация продления": "iteration_0",
  "Дата вывода 1 - итерация продления": "iteration_1",
  "Дата вывода 1 - итерация продления (корр 23.06.22)": "iteration_1_1",
  "Дата вывода 2 - итерация продления (корр 30.06.22)": "iteration_2"
}
iterations_list = ["operation_finish_date_iteration_0",	"operation_finish_date_iteration_1",	"operation_finish_date_iteration_2"]
iterations_dict = {
  "operation_finish_date_iteration_0":"Дата вывода 0 - итерация продления",
  "operation_finish_date_iteration_1":"Дата вывода 1 - итерация продления",
  "operation_finish_date_iteration_2":"Дата вывода 2 - итерация продления"
}


def read_be_2_eo_xlsx():
  # читаем загруженный файл 
  be_eo_raw_data = pd.read_excel('uploads/be_eo_2_data.xlsx', sheet_name='be_eo_data', index_col = False)
  be_eo_data = be_eo_raw_data.rename(columns=be_data_columns_to_master_columns)

  be_eo_data['eo_code'] = be_eo_data['eo_code'].astype(str)
  be_eo_data['operation_finish_date_iteration_0'] = pd.to_datetime(be_eo_data['operation_finish_date_iteration_0'], format='%d.%m.%Y')
  be_eo_data['operation_finish_date_iteration_1'] = pd.to_datetime(be_eo_data['operation_finish_date_iteration_1'], format='%d.%m.%Y')
  be_eo_data['operation_finish_date_iteration_2'] = pd.to_datetime(be_eo_data['operation_finish_date_iteration_2'], format='%d.%m.%Y')
  be_eo_data['conservation_start_date'] = pd.to_datetime(be_eo_data['conservation_start_date'], format='%d.%m.%Y')

  # получаем данные из мастер-файла
  con = sqlite3.connect("database/datab.db")
  cursor = con.cursor()
  # sql = "SELECT * FROM eo_DB JOIN be_DB"
  sql = "SELECT \
  eo_DB.eo_code, \
  eo_DB.temp_eo_code, \
  eo_DB.temp_eo_code_status, \
  eo_DB.be_code, \
  eo_DB.head_type, \
  be_DB.be_description, \
  eo_DB.eo_class_code, \
  eo_class_DB.eo_class_description, \
  models_DB.eo_model_name, \
  models_DB.eo_category_spec, \
  models_DB.type_tehniki, \
  models_DB.marka_oborudovania, \
  eo_DB.eo_model_id, \
  eo_DB.eo_description, \
  eo_DB.gar_no, \
  eo_DB.constr_type, \
  eo_DB.constr_type_descr, \
  eo_DB.operation_start_date, \
  eo_DB.expected_operation_period_years, \
  eo_DB.operation_finish_date_calc, \
  eo_DB.sap_planned_finish_operation_date, \
  eo_DB.operation_finish_date_sap_upd, \
  eo_DB.sap_system_status, \
  eo_DB.sap_user_status, \
  eo_DB.prodlenie_2022 \
  FROM eo_DB \
  LEFT JOIN models_DB ON eo_DB.eo_model_id = models_DB.eo_model_id \
  LEFT JOIN be_DB ON eo_DB.be_code = be_DB.be_code \
  LEFT JOIN eo_class_DB ON eo_DB.eo_class_code = eo_class_DB.eo_class_code \
  LEFT JOIN operation_statusDB ON eo_DB.expected_operation_status_code = operation_statusDB.operation_status_code"
  
  master_eo_df = pd.read_sql_query(sql, con)
  # подменяем временнные номера
  master_eo_df_temp_eo = master_eo_df.loc[master_eo_df['temp_eo_code_status']=='temp_eo_code']
  indexes_temp_eo = list(master_eo_df_temp_eo.index.values)
  # print(master_eo_df_temp_eo['temp_eo_code'])
  # print(list(master_eo_df_temp_eo.loc[indexes_temp_eo, ['temp_eo_code']]))  
  master_eo_df.loc[indexes_temp_eo, ['eo_code']] = master_eo_df_temp_eo['temp_eo_code']
  # print(master_eo_df.loc[indexes_temp_eo, ['eo_code']])
  # print('должен быть список', master_eo_df.loc[indexes_temp_eo, ['eo_code']])
  
  master_eo_df = master_eo_df.loc[master_eo_df['head_type']=='head']
  master_eo_df['operation_start_date'] = pd.to_datetime(master_eo_df['operation_start_date'])
  master_eo_df['sap_planned_finish_operation_date'] = pd.to_datetime(master_eo_df['sap_planned_finish_operation_date'])
  master_eo_df['operation_finish_date_sap_upd'] = pd.to_datetime(master_eo_df['operation_finish_date_sap_upd'])
  # джойним данные из файла с мастер-данными
  be_master_data = pd.merge(master_eo_df, be_eo_data, on='eo_code', how='left')


  
  result_data_list = []
  

  # итерируемся по списку итераций
  for iteration, iteration_rus in iterations_dict.items():
    be_master_data[iteration] = pd.to_datetime(be_master_data[iteration])
    be_master_data[iteration].fillna(date_time_plug, inplace = True)
    be_master_data['operation_finish_date'] = be_master_data['operation_finish_date_sap_upd']
    be_master_data['operation_finish_date'] = be_master_data['operation_finish_date'].dt.date
    
    be_master_data[iteration] = be_master_data[iteration].dt.date
    
    be_master_data_temp = be_master_data.loc[be_master_data[iteration]!=date_time_plug.date()]

    indexes = list(be_master_data_temp.index.values)
    
    be_master_data.loc[indexes, ['operation_finish_date']] = be_master_data_temp[iteration]

    be_master_data['operation_finish_date'] = pd.to_datetime(be_master_data['operation_finish_date'])
    
    # итерируемся по списку ео в загруженном файле
    i=0
    lenght = len(be_master_data)
    print("Итерация ", iteration,", ", lenght, " eo")
    for row in be_master_data.itertuples():
      i=i+1
      # print("Итерация ", iteration,", ", i, " из ", lenght)
      eo_code = getattr(row, 'eo_code')
      be_code = getattr(row, 'be_code')
      be_description = getattr(row, 'be_description')
      eo_class_code = getattr(row, 'eo_class_code')
      eo_class_description = getattr(row, 'eo_class_description')
      eo_model_name = getattr(row, 'eo_model_name')
      # eo_model_id = getattr(row, 'eo_model_id')
      eo_category_spec = getattr(row, 'eo_category_spec')
      type_tehniki = getattr(row, 'type_tehniki')
      marka_oborudovania = getattr(row, 'marka_oborudovania')
      
      eo_description = getattr(row, "eo_description")
      gar_no = getattr(row, "gar_no")
      constr_type = getattr(row, "constr_type")
      constr_type_descr = getattr(row, "constr_type_descr")
      # operation_status_rus = getattr(row, "operation_status")
      sap_user_status = getattr(row, "sap_user_status")
      sap_system_status = getattr(row, "sap_system_status")
      operation_status_from_file = getattr(row, "operation_status") # статус, полученный из файла

      operation_start_date = getattr(row, 'operation_start_date')
      expected_operation_period_years = getattr(row, 'expected_operation_period_years')
      operation_finish_date_calc = getattr(row, 'operation_finish_date_calc')
      sap_planned_finish_operation_date = getattr(row, 'sap_planned_finish_operation_date')
      operation_finish_date_sap_upd = getattr(row, 'operation_finish_date_sap_upd')
      operation_finish_date_update_iteration = getattr(row, iteration)
      operation_finish_date = getattr(row, 'operation_finish_date')
      iteration_name = iteration_rus
      conservation_start_date = getattr(row, 'conservation_start_date')
      
      status_condition_dict = {
        "new":"Ввод нового",
        "on_balance":"На балансе",
        "conservation":"Консервация",
        "remake":"Переоборудование",
        "out":"План на вывод",
        "in_operation":"Эксплуатация"
      }
      for status_condition, status_condition_rus in status_condition_dict.items():
        for year, year_data in year_dict.items():
          # print(year)
          year_first_date = datetime.strptime(year_data['period_start'], '%d.%m.%Y')
          year_last_date = datetime.strptime(year_data['period_end'], '%d.%m.%Y')
          temp_dict = {} 
          temp_dict['eo_code'] = eo_code
          temp_dict['be_code'] = be_code
          temp_dict['be_description'] = be_description
          temp_dict['eo_class_code'] = eo_class_code
          temp_dict['eo_class_description'] = eo_class_description
          # temp_dict['eo_model_id'] = eo_model_id
          temp_dict['eo_category_spec'] = eo_category_spec
          temp_dict['eo_model_name'] = eo_model_name
          temp_dict['type_tehniki'] = type_tehniki
          temp_dict['marka_oborudovania'] = marka_oborudovania
          
          temp_dict['eo_description'] = eo_description
          temp_dict['gar_no'] = gar_no
          temp_dict['constr_type'] = constr_type
          temp_dict['constr_type_descr'] = constr_type_descr
          
          temp_dict['sap_user_status'] = sap_user_status
          temp_dict['sap_system_status'] = sap_system_status
          temp_dict['operation_start_date'] = operation_start_date
          temp_dict['expected_operation_period_years'] = expected_operation_period_years
          temp_dict['operation_finish_date_calc'] = operation_finish_date_calc
          temp_dict['sap_planned_finish_operation_date'] = sap_planned_finish_operation_date
          temp_dict['operation_finish_date_sap_upd'] = operation_finish_date_sap_upd
          temp_dict['operation_finish_date_update_iteration'] = operation_finish_date_update_iteration
          temp_dict['operation_finish_date'] = operation_finish_date
          temp_dict['operation_status_from_file'] = operation_status_from_file
          temp_dict['conservation_start_date'] = conservation_start_date
          temp_dict['iteration_name'] = iteration_name
          temp_dict['year'] = year 

          
              	

          if status_condition == "new":
            temp_dict['operation_status'] = "Ввод нового"
            if sap_user_status not in sap_user_status_cons_status_list and \
            sap_system_status not in sap_system_status_ban_list and \
            operation_start_date >= year_first_date and \
            operation_start_date <= year_last_date:
              temp_dict['qty'] = 1
              temp_dict['Ввод нового'] = 1
              temp_dict['Возраст'] = (year_last_date - operation_start_date).days / 365.25
              result_data_list.append(temp_dict)
            # else:
            #   temp_dict['Ввод нового'] = 0
            #   temp_dict['Возраст'] = 0
            #  temp_dict['qty'] = 0
            # result_data_list.append(temp_dict)  
          if status_condition == "in_operation":
            temp_dict['operation_status'] = "Эксплуатация"
            if sap_user_status not in sap_user_status_cons_status_list and \
            operation_status_from_file != "Консервация" and \
            sap_system_status not in sap_system_status_ban_list and \
            operation_start_date <= year_last_date and \
            operation_finish_date >= year_first_date:
              temp_dict['qty'] = 1
              temp_dict['Эксплуатация'] = 1
              temp_dict['Возраст'] = (year_last_date - operation_start_date).days / 365.25
              result_data_list.append(temp_dict)  
            # else:
            #   temp_dict['Эксплуатация'] = 0
            #   temp_dict['Возраст'] = 0
            #   temp_dict['qty'] = 0
            # result_data_list.append(temp_dict)
            
          if status_condition == "conservation":
            temp_dict['operation_status'] = "Консервация"
            
            if operation_status_from_file == "Консервация" and \
             sap_system_status not in sap_system_status_ban_list and \
             conservation_start_date <= year_last_date and \
             operation_finish_date >= year_first_date:
             temp_dict['qty'] = 1
             temp_dict['Консервация'] = 1

             result_data_list.append(temp_dict)  
            # else:
            #   temp_dict['Консервация'] = 0
            #   temp_dict['Возраст'] = 0
            #   temp_dict['qty'] = 0
            # result_data_list.append(temp_dict)
          
          if status_condition == "on_balance":
            temp_dict['operation_status'] = "На балансе"
            
            if (operation_status_from_file == "Консервация" and \
             sap_system_status not in sap_system_status_ban_list and \
             conservation_start_date <= year_last_date and \
             operation_finish_date >= year_first_date) or \
             (sap_user_status not in sap_user_status_cons_status_list and \
             sap_system_status not in sap_system_status_ban_list and \
             operation_start_date <= year_last_date and \
             operation_finish_date >= year_first_date):
               temp_dict['qty'] = 1
               temp_dict['На балансе'] = 1
               result_data_list.append(temp_dict)  
            # else:
            #   temp_dict['На балансе'] = 0
            #   temp_dict['Возраст'] = 0
            #   temp_dict['qty'] = 0
            # result_data_list.append(temp_dict)

          if status_condition == "out":
            temp_dict['operation_status'] = "План на вывод"
            if sap_system_status not in sap_system_status_ban_list and \
            operation_finish_date <= year_last_date and \
            operation_finish_date >= year_first_date:
              temp_dict['qty'] = -1
              temp_dict['План на вывод'] = -1
              temp_dict['Возраст'] = (operation_finish_date - operation_start_date).days / 365.25
              result_data_list.append(temp_dict)  
            # else:
            #   temp_dict['План на вывод'] = 0
            #   temp_dict['Возраст'] = 0
            #   temp_dict['qty'] = 0
            # result_data_list.append(temp_dict)

          # if status_condition == "in_operation" and iteration == 'operation_finish_date_iteration_2' and year == 2023:
          #   # получаем строки с данными по капремонтам
          #   if eo_code =='100000028868':
          #     krik_data_temp = krik_data.loc[krik_data['year']==year]
          #     krik_data_temp = krik_data_temp.loc[krik_data_temp['eo_code']==eo_code]
            
          #     # print(krik_data_temp)
          #     # print("status_condition: ", status_condition, " iteration:", iteration, " year: ", year)
            
          #     if len(krik_data_temp) >0:
          #       print("длина списка ремонтов 100000028868: ", len(krik_data_temp))
          #       i=0
          #       for row in krik_data_temp.itertuples():
          #         i = i+1
          #         print("запись ремонта ", i, " из ", len(krik_data_temp))
          #         overhaul_type = getattr(row, 'overhaul_type')
                  
          #         print("eo_code:", eo_code, "overhaul_type:", overhaul_type)
          #         temp_dict['overhaul_type'] = overhaul_type
          #         # print(temp_dict)
          #         print("длина общего списка: ", len(result_data_list))
          #         result_data_list.append(temp_dict)
          #         print(result_data_list[-1])
       
          #       temp_df = pd.DataFrame(result_data_list)
          #       print(temp_df.head(-10))
          #       temp_df.to_csv('temp_data/temp_df_check.csv')
              
            #     if eo_code =='100000028868':
            #       print(overhaul_type)
            #     overhaul_plan_date = getattr(row, 'overhaul_plan_date')
            #     counter = getattr(row, 'counter')
            #     Cost_rub = getattr(row, 'Cost_rub')
            #     temp_dict['qty'] = 0
            #     temp_dict['Эксплуатация'] = 0
            #     temp_dict['На балансе'] = 0
            #     temp_dict['Возраст'] = 0
            #     temp_dict['overhaul_type'] = overhaul_type
            #     temp_dict['overhaul_plan_date'] = overhaul_plan_date
            #     temp_dict['counter'] = counter
            #     temp_dict['Cost_rub'] = Cost_rub
            #     result_data_list.append(temp_dict)

  iterations_df = pd.DataFrame(result_data_list) 
  try:
    iterations_df['be_code'] = iterations_df['be_code'].astype(int)
  except:
    pass
  try:
    iterations_df["operation_start_date"] = iterations_df["operation_start_date"].dt.strftime("%d.%m.%Y")
  except:
    pass
  try:
    iterations_df["operation_finish_date_calc"] = pd.to_datetime(iterations_df["operation_finish_date_calc"])
    iterations_df["operation_finish_date_calc"] = iterations_df["operation_finish_date_calc"].dt.strftime("%d.%m.%Y")
  except Exception as e:
    print(e)  
  try:
    iterations_df["sap_planned_finish_operation_date"] = pd.to_datetime(iterations_df["sap_planned_finish_operation_date"])
    iterations_df["sap_planned_finish_operation_date"] = iterations_df["sap_planned_finish_operation_date"].dt.strftime("%d.%m.%Y")
  except Exception as e:
    print(e)
  try:
    iterations_df["operation_finish_date_sap_upd"] = pd.to_datetime(iterations_df["operation_finish_date_sap_upd"])
    iterations_df["operation_finish_date_sap_upd"] = iterations_df["operation_finish_date_sap_upd"].dt.strftime("%d.%m.%Y")
  except Exception as e:
    print(e)
  try:
    iterations_df["operation_finish_date_update_iteration"] = pd.to_datetime(iterations_df["operation_finish_date_update_iteration"])
    iterations_df["operation_finish_date_update_iteration"] = iterations_df["operation_finish_date_update_iteration"].dt.date
    iterations_df_temp = iterations_df.loc[iterations_df["operation_finish_date_update_iteration"]==date_time_plug.date()]
    indexes = list(iterations_df_temp.index.values)
    iterations_df.loc[indexes, ['operation_finish_date_update_iteration']] = ""
    iterations_df["operation_finish_date_update_iteration"] = pd.to_datetime(iterations_df["operation_finish_date_update_iteration"])
    iterations_df["operation_finish_date_update_iteration"] = iterations_df["operation_finish_date_update_iteration"].dt.strftime("%d.%m.%Y")
  except Exception as e:
    print(e)  

  try:
    iterations_df["operation_finish_date"] = pd.to_datetime(iterations_df["operation_finish_date"])
    iterations_df["operation_finish_date"] = iterations_df["operation_finish_date"].dt.strftime("%d.%m.%Y")
  except Exception as e:
    print(e)

  try:
    iterations_df["conservation_start_date"] = pd.to_datetime(iterations_df["conservation_start_date"])
    iterations_df["conservation_start_date"] = iterations_df["conservation_start_date"].dt.strftime("%d.%m.%Y")
  except Exception as e:
    print(e)

  # читаем файл с данными КРИК
  krik_data = pd.read_excel('temp_data/krik.xlsx')
  krik_data['eo_code'] = krik_data['eo_code'].astype(str)
  krik_data['overhaul_plan_date'] = pd.to_datetime(krik_data['overhaul_plan_date'], format='%d.%m.%Y')
  krik_data['year'] = krik_data['overhaul_plan_date'].dt.year
  krik_data['iteration_name'] = iterations_dict[iterations_list[-1]]
  krik_data['год капремонта'] = krik_data['overhaul_plan_date'].dt.year
  krik_data['operation_status'] = 'Эксплуатация'
  

  
  be_master_data_temp = be_master_data.loc[:, ['eo_code', 'be_code', 'be_description', 'eo_class_code', 'eo_class_description', 'eo_category_spec', 'eo_model_name', 'type_tehniki', 'marka_oborudovania', 'eo_description', 'gar_no', 'constr_type', 'constr_type_descr', 'sap_user_status', 'sap_system_status', 'operation_start_date', 'expected_operation_period_years', 'operation_finish_date_calc', 'sap_planned_finish_operation_date', 'operation_finish_date_sap_upd', 'operation_finish_date', 'prodlenie_2022']]
  

  krik_data_eo = pd.merge(krik_data, be_master_data_temp, on='eo_code')
  krik_data_eo = krik_data_eo.rename(columns=master_data_to_ru_columns)
  print("start krik_data_eo excel")
  wb = openpyxl.Workbook(write_only=True)
  ws = wb.create_sheet("капремонты")
  for r in dataframe_to_rows(krik_data_eo, index=True):
    i=i+1
    # print(i, " из ", lenght)
    ws.append(r)

  wb.save("temp_data/krik_data.xlsx")
  print("finish krik_data_eo excel")
  
  iterations_and_overhaul_df = pd.concat([iterations_df, krik_data_eo])
  
  iterations_and_overhaul_df.to_csv('temp_data/iterations_and_overhaul_df.csv')
    
  iterations_df_rus = iterations_and_overhaul_df.rename(columns=master_data_to_ru_columns)
  # iterations_df.to_csv('temp_data/iterations_df.csv')
  # iterations_df.to_excel('temp_data/iterations_df.xlsx')
  print("start full_data_eo excel")  
  wb = openpyxl.Workbook(write_only=True)
  ws = wb.create_sheet("Sheet1")
  
  i=0
  iterations_df_rus.reset_index()
  lenght = len(iterations_df_rus)
  print(len(iterations_df_rus))
  
  for r in dataframe_to_rows(iterations_df_rus, index=True):
    i=i+1
    # print(i, " из ", lenght)
    ws.append(r)

  wb.save("temp_data/iterations_df.xlsx")
  print("finish full_data_eo excel") 
  
  
  
  # итерируемся по итоговому датафрейму
  # из итогового датафрейма отбираем строки
  # last_iteration  = list(iterations_dict.keys())[-1]
  # print(last_iteration)
  # iterations_df_partial = iterations_df.loc[iterations_df['Итерация продления'] == last_iteration]
  # iterations_df_partial = iterations_df_partial.loc[iterations_df_partial['year']==2023]
  # iterations_eo_list = list(set(iterations_df['eo_code']))
  # result_df_list = []
  # for eo in iterations_eo_list:
    
    #получаем записи из крик файла
    # if eo == '100000028868':
  #   krik_data_partial = krik_data.loc[krik_data['eo_code']==eo]
  #   krik_data_partial = krik_data_partial.loc[krik_data_partial['year']==2023]
  #   for row in krik_data_partial.itertuples():
  #     temp_dict = {}
  #     overhaul_type = getattr(row, 'overhaul_type')
  #     overhaul_plan_date = getattr(row, 'overhaul_plan_date')
  #     overhaul_counter = getattr(row, 'counter')
  #     cost_rub = getattr(row, 'Cost_rub')
  #     year = getattr(row, 'year')
  #     temp_dict['eo_code'] = eo
  #     temp_dict['overhaul_type'] = overhaul_type
  #     print(overhaul_type)
  #     temp_dict['overhaul_plan_date'] = overhaul_plan_date
  #     temp_dict['overhaul_counter'] = overhaul_counter
  #     temp_dict['cost_rub'] = cost_rub
  #     temp_dict['year'] = year
  #     result_df_list.append(temp_dict)

  # overhaul_df = pd.DataFrame(result_df_list)
  # print(overhaul_df)
  # iterations_and_overhaul_df = pd.concat([iterations_df, overhaul_df])
  # iterations_and_overhaul_df.to_csv('temp_data/iterations_and_overhaul_df.csv')
  
  
  # wb = openpyxl.Workbook(write_only=True)
  # ws = wb.create_sheet("Sheet1")
  
  # i=0
  # iterations_df.reset_index()
  # iterations_df_1 = iterations_df.iloc[:100000,:]
  # lenght = len(iterations_df_1)
  # print(len(iterations_df_1))
  
  # for r in dataframe_to_rows(iterations_df_1, index=True):
  #   i=i+1
  #   print(i, " из ", lenght)
  #   ws.append(r)

  # # Now save
  # wb.save("temp_data/iterations_df_1.xlsx")
  # # print('start excel')
  # # iterations_df.to_excel("temp_data/iterations.xlsx")


  # ##############################################################################################
  # wb = openpyxl.Workbook(write_only=True)
  # ws = wb.create_sheet("Sheet1")
        
  # i=0
  # iterations_df_2 = iterations_df.iloc[100000:200000,:]
  # lenght = len(iterations_df_2)
  # print(len(iterations_df_2))
  
  # for r in dataframe_to_rows(iterations_df_2, index=True):
  #   i=i+1
  #   print(i, " из ", lenght)
  #   ws.append(r)

  # # Now save
  # wb.save("temp_data/iterations_df_2.xlsx")

  #  ##############################################################################################
  # wb = openpyxl.Workbook(write_only=True)
  # ws = wb.create_sheet("Sheet1")
        
  # i=0
  # iterations_df_3 = iterations_df.iloc[200000:,:]
  # lenght = len(iterations_df_3)
  # print(len(iterations_df_3))
  
  # for r in dataframe_to_rows(iterations_df_3, index=True):
  #   i=i+1
  #   print(i, " из ", lenght)
  #   ws.append(r)

  # # Now save
  # wb.save("temp_data/iterations_df_3.xlsx")

        
        